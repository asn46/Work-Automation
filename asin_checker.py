import os
import PySimpleGUI as sg
from playwright.sync_api import sync_playwright
from datetime import date
from time import sleep as timeSleep
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter as gcl

#-----------------------------------------------------------------------


file = r'C:\Users\anerl\Desktop\Amazon ASIN search.xlsx'

wb =  load_workbook(file)
wb_sheet = wb['Sheet2']
max_row = wb_sheet.max_row

with sync_playwright() as p:
    browser = p.chromium.launch(headless=False)
    context = context = browser.new_context()
    
    page = context.new_page()
    page.goto('https://www.amazon.com/')
    page.is_visible('div.nav-search-field')

    for row_num in range(2, max_row+1):
        ASIN = wb_sheet[f'A{row_num}'].value
        name = wb_sheet[f'C{row_num}'].value
        
        page.fill('input#twotabsearchtextbox', ASIN)
        page.click('input#nav-search-submit-button')
        
        timeSleep(1)
        
        try:
            if page.is_visible(f'text={name}'):
                wb_sheet[f'D{row_num}'].value = 'Yes'
                
            else:
                wb_sheet[f'D{row_num}'].value = 'No'
        except: 
            print('There was an error')
    
    wb.save(r'C:\Users\anerl\Desktop\Amazon ASIN search.xlsx')
            
context.close()
browser.close()
        
#---------------------------------------------------------------------




